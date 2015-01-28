# coding: utf-8

# my_utils for Intro to Data Science with Python
# Author: Kat Chuang
# Created: Nov 2014

# --------------------------------------
## Stage 2 begin

import csv 

# open a file and return a double list
def open_with_CSV(filename, d='\t'):
    uuids = []
    with open(filename) as tsvin:
        tsvin = csv.reader(tsvin, delimiter=d)
        for row in tsvin:
            uuids.append(row)
    return uuids

#2.a size
def number_of_records(data_sample):
    return len(data_sample)


#2.b calculate sum
def calculate_sum(data_sample):
    total = 0
    for row in data_sample[1:]:
        price = float(row[2])
        total += price
    return total

#2.c 1 # Find the average price
def find_average(data_sample, headers=False):
    total = calculate_sum(data_sample)
    size = number_of_records(data_sample)
    if headers:
        total -= 1

    average = total / size
    return average

#2.d. Max, Min
# Find the maximum price
def find_max(theData, col):
    tempList = []
    
    for row in theData:
        price = float(row[col])
        tempList.append(price)
    return max(tempList)

# Find the minimum price
def find_min(theData, col):
    tempList = []

    for row in theData:
        price = float(row[col])
        tempList.append(price)
    return min(tempList)

# Find the max or min price
def find_max_min(the_data, col, m):
    tempList = [], val = 0

    for row in the_data:
        price = float(row[col])
        tempList.append(price)

        if m == "max": 
            val = max(tempList)
        elif m == "min":
            val = min(tempList)
        else: # hopefully we don’t come to this
            pass 

    return val


## Stage 2 end

# --------------------------------------

## Stage 3 begin
# Stage 3: Cleaning data

# Filter rows were columns match a string data type
def filter_col_by_string(the_data, field, filter_condition):
    filtered_rows = []
    
    #find index of field in first row
    col = int(the_data[0].index(field))
    filtered_rows.append(the_data[0])

    for row in the_data[1:]:
        if row[col] == filter_condition:
            filtered_rows.append(row)
            
    return filtered_rows

# Filter rows were columns match a float data type
def filter_col_by_float(the_data, field, direction, filter_condition):
    filtered_rows = []
    
    #find index of field in first row
    col = int(the_data[0].index(field))
    cond = float(filter_condition)
    
    for row in the_data[1:]:
        element = float(row[col])
        
        if direction == "<":
            if element < cond: filtered_rows.append(row)
                
        elif direction == "<=":
            if element <= cond: filtered_rows.append(row)

        elif direction == ">":
            if element > cond: filtered_rows.append(row)

        elif  direction == ">=":
            if element >= cond: filtered_rows.append(row)
                
        elif  direction == "==":
            if element == cond: filtered_rows.append(row)
        else:
            pass
        
    return filtered_rows


## Stage 3 end

# --------------------------------------

## Stage 4 begin

#4.a csv
def write_to_file(filename, data_sample):
    example = csv.writer(open(filename, 'w', newline=''))
    example.writerows(data_sample)

#4.b more functions
def write_brand_and_price_to_file(filename, data_sample):

    # confirm that the columns only have two columns, otherwise take the two fields
    num_fields = len(data_sample[0])
    
    brand_field_index = 5 #int(dataSample[0].index("brand"))
    price_field_index = 2 #int(dataSample[0].index("priceLabel"))
    
    #if numFields > 2:
    new_array = []
    for record in data_sample:
        new_record = [None] * 2
        new_record[0] = record[brand_field_index]
        new_record[1] = record[price_field_index]
        new_array.append(new_record)

    # write the file
    write_to_file(filename, new_array)  

#4.c export to excel
from openpyxl import Workbook
from openpyxl.cell import get_column_letter

def save_spreadsheet(filename, data_sample):
    
    wb = Workbook()
    ws = wb.active

    rowIndex = 1
    for rows in data_sample:
        colIndex = 1 
        for field in rows:
            colIndex2 = get_column_letter(colIndex)
            ws.cell('%s%s'%(colIndex2, rowIndex)).value = field
            colIndex +=1
        rowIndex += 1
    
    wb.save(filename)

## Stage 4 end

# --------------------------------------

## Stage 5 begin
#5.a Line charts
def create_line_chart(sample, title, exported_figure_filename):
    fig = plt.figure()
    ax = fig.add_subplot(1, 1, 1)

    prices = sorted(map(int, sample))
    x_axis_ticks = list( range(len(sample)) )
    ax.plot(x_axis_ticks, prices, label='price points', linewidth=2)
    ax.set_title(title)
    ax.set_xlabel('Tie Price ($)')
    ax.set_ylabel('Number of Ties')
    ax.set_xlim([0,len(sample)])
    fig.savefig(exported_figure_filename)


#5.b bar charts
def create_bar_chart(price_groups, exported_figure_filename):
    fig = plt.figure()
    ax = fig.add_subplot(1, 1, 1)
    colors=plt.rcParams['axes.color_cycle']
     
    for group in price_groups:
        ax.bar(group, price_groups[group], color=colors[group%len(price_groups)])

    labels = ["$0-50", "$50-100", "$100-150", "$150-200", "$200-250", "$250+"]
    ax.legend(labels)

    ax.set_title('Amount of Ties at price points')
    ax.set_xlabel('Tie Price ($)')
    ax.set_xticklabels(labels, ha='left')
    ax.set_xticks( range(1, len(price_groups)+1) )
    ax.set_ylabel('Number of Ties')

    plt.grid(True)
    fig.savefig(exported_figure_filename)


def print_brand_avg_min(name):
    tieSample = filterByString(dataFromCSV, "brandName", name)
    avgPrice = calculateSum(tieSample) / len(tieSample)
    minPrice = findMin(tieSample[1:], 2)
    print("{2} Average: ${0:6.2f}; Min: ${1:.2f}".format(avgPrice, minPrice, name))
#5.c tables
from collections import Counter
def group_prices_by_range(prices_in_float):
    
    tally = Counter()

    for item in prices_in_float:
        bucket = 0
        rounded_price = round(item, -1)
        if rounded_price >= 0 and rounded_price <= 50:
            bucket = 1
        elif rounded_price >= 50 and rounded_price <= 100:
            bucket = 2
        elif rounded_price >= 100 and rounded_price <= 150:
            bucket = 3
        elif rounded_price >= 150 and rounded_price <= 200:
            bucket = 4
        elif rounded_price >= 200 and rounded_price <= 250:
            bucket = 5
        elif rounded_price >= 250:
            bucket = 6
        else:
            bucket = 7

        tally[bucket] += 1
    return tally


## Stage 5 end

# --------------------------------------

## Stage 6 begin

## Stage 6 end


